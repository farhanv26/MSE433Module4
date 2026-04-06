"""AI Event Log export (CSV + Excel): seven columns, all segments, chronological."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Human-readable headers matching the AI Event Log layout.
EVENT_LOG_COLUMNS = [
    "Start Time",
    "End Time",
    "Duration (s)",
    "Procedure Phase",
    "Delay Category",
    "Description",
    "Confidence",
]


def _sort_key(seg: Dict) -> Any:
    t = seg.get("start_time")
    if t is None:
        return 0.0
    try:
        return float(t)
    except (TypeError, ValueError):
        return 0.0


def build_event_log_dataframe(classified_segments: List[Dict]) -> pd.DataFrame:
    """Full event log: every segment, sorted by start time. Empty input → valid empty schema."""
    ordered = sorted(classified_segments, key=_sort_key)
    rows: List[Dict] = []
    for seg in ordered:
        rows.append(
            {
                "Start Time": seg.get("start_time"),
                "End Time": seg.get("end_time"),
                "Duration (s)": seg.get("duration_sec"),
                "Procedure Phase": seg.get("procedure_phase", "Unknown"),
                "Delay Category": seg.get("delay_category", "-"),
                "Description": seg.get("description", ""),
                "Confidence": seg.get("confidence", 0.0),
            }
        )
    return pd.DataFrame(rows, columns=EVENT_LOG_COLUMNS)


def _format_ai_event_log_sheet(ws) -> None:
    """Bold header row, freeze header, auto-fit column widths (openpyxl-only, no data changes)."""
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val).strip()))
        # Readable minimum for headers like "Procedure Phase"; cap very long description cells
        width = min(max(max_len + 2, 12), 72)
        ws.column_dimensions[letter].width = width


def export_event_log(
    events_df: pd.DataFrame,
    csv_path: str | Path,
    xlsx_path: str | Path,
    *,
    sheet_name: str = "AI Event Log",
) -> None:
    csv_path, xlsx_path = Path(csv_path), Path(xlsx_path)
    if list(events_df.columns) != EVENT_LOG_COLUMNS:
        events_df = events_df.reindex(columns=EVENT_LOG_COLUMNS)
    events_df.to_csv(csv_path, index=False)
    safe_name = sheet_name[:31]
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        events_df.to_excel(writer, index=False, sheet_name=safe_name)
        _format_ai_event_log_sheet(writer.book[safe_name])
    logger.info("Saved event log CSV/XLSX: %s, %s", csv_path, xlsx_path)
